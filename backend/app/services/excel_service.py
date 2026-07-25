"""
Weekly Excel report generation using openpyxl.
Generates formatted .xlsx with new contacts and enquiries, with deduplication.
"""

import io
import logging
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from ..models.db import Contact, Enquiry, WeeklyReportLog

logger = logging.getLogger(__name__)

# ── Style constants ────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=14)
HEADER_FILL = PatternFill(start_color="BF953F", end_color="BF953F", fill_type="solid")
ROW_FILL_A = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
ROW_FILL_B = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D5C4A1"),
    right=Side(style="thin", color="D5C4A1"),
    top=Side(style="thin", color="D5C4A1"),
    bottom=Side(style="thin", color="D5C4A1"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(ws, headers, row=1):
    """Apply header styling to a worksheet row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


def _style_data_row(ws, row_idx, col_count):
    """Apply alternating row styling."""
    fill = ROW_FILL_A if row_idx % 2 == 0 else ROW_FILL_B
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = BODY_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = LEFT_ALIGN


async def generate_weekly_report(db: AsyncSession) -> tuple[io.BytesIO | None, int]:
    """
    Generate an Excel report with new contacts and enquiries since the last report.
    
    Returns (buffer, record_count) — buffer is None if no new records.
    """
    now = datetime.now(timezone.utc)
    one_week_ago = now - timedelta(days=7)

    # ── Find last report to get max IDs for deduplication ───
    last_report = await db.execute(
        select(WeeklyReportLog).order_by(WeeklyReportLog.sent_at.desc()).limit(1)
    )
    last = last_report.scalar_one_or_none()
    last_contact_id = last.max_contact_id if last else 0
    last_enquiry_id = last.max_enquiry_id if last else 0

    # ── Fetch new contacts ─────────────────────────────────
    contacts_result = await db.execute(
        select(Contact)
        .where(Contact.id > last_contact_id)
        .where(Contact.created_at >= one_week_ago)
        .order_by(Contact.created_at)
    )
    contacts = contacts_result.scalars().all()

    # ── Fetch new enquiries ────────────────────────────────
    enquiries_result = await db.execute(
        select(Enquiry)
        .where(Enquiry.id > last_enquiry_id)
        .where(Enquiry.created_at >= one_week_ago)
        .order_by(Enquiry.created_at)
    )
    enquiries = enquiries_result.scalars().all()

    total_records = len(contacts) + len(enquiries)
    if total_records == 0:
        logger.info("No new records for weekly report — skipping.")
        return None, 0

    # ── Build workbook ─────────────────────────────────────
    wb = Workbook()

    # Sheet 1: Contacts
    ws_contacts = wb.active
    ws_contacts.title = "New Contacts"
    contact_headers = ["#", "Name", "Email", "Phone", "Subject", "Message", "Location", "Date"]
    _style_header(ws_contacts, contact_headers)

    for idx, c in enumerate(contacts, 1):
        row = idx + 1
        location = c.location_name or ""
        if c.latitude and c.longitude:
            location += f" ({c.latitude}, {c.longitude})"
        ws_contacts.cell(row=row, column=1, value=idx)
        ws_contacts.cell(row=row, column=2, value=f"{c.first_name} {c.last_name}")
        ws_contacts.cell(row=row, column=3, value=c.email)
        ws_contacts.cell(row=row, column=4, value=c.phone)
        ws_contacts.cell(row=row, column=5, value=c.subject)
        ws_contacts.cell(row=row, column=6, value=c.message[:300])
        ws_contacts.cell(row=row, column=7, value=location.strip())
        ws_contacts.cell(row=row, column=8, value=c.created_at.strftime("%d %b %Y, %I:%M %p") if c.created_at else "")
        _style_data_row(ws_contacts, row, len(contact_headers))

    # Auto-fit column widths
    for col_idx in range(1, len(contact_headers) + 1):
        ws_contacts.column_dimensions[chr(64 + col_idx)].width = max(18, len(contact_headers[col_idx - 1]) + 6)
    ws_contacts.column_dimensions["F"].width = 50  # Message column wider
    ws_contacts.column_dimensions["G"].width = 35  # Location column wider

    # Sheet 2: Enquiries
    ws_enquiries = wb.create_sheet(title="New Enquiries")
    enquiry_headers = ["#", "Ref", "Name", "Email", "Phone", "Service", "Message", "Location", "Date"]
    _style_header(ws_enquiries, enquiry_headers)

    for idx, e in enumerate(enquiries, 1):
        row = idx + 1
        location = e.location_name or ""
        if e.latitude and e.longitude:
            location += f" ({e.latitude}, {e.longitude})"
        ws_enquiries.cell(row=row, column=1, value=idx)
        ws_enquiries.cell(row=row, column=2, value=e.reference)
        ws_enquiries.cell(row=row, column=3, value=e.name)
        ws_enquiries.cell(row=row, column=4, value=e.email)
        ws_enquiries.cell(row=row, column=5, value=e.phone)
        ws_enquiries.cell(row=row, column=6, value=e.service)
        ws_enquiries.cell(row=row, column=7, value=e.message[:300])
        ws_enquiries.cell(row=row, column=8, value=location.strip())
        ws_enquiries.cell(row=row, column=9, value=e.created_at.strftime("%d %b %Y, %I:%M %p") if e.created_at else "")
        _style_data_row(ws_enquiries, row, len(enquiry_headers))

    for col_idx in range(1, len(enquiry_headers) + 1):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(65 + (col_idx - 1) % 26)
        ws_enquiries.column_dimensions[col_letter].width = max(18, len(enquiry_headers[col_idx - 1]) + 6)
    ws_enquiries.column_dimensions["G"].width = 50
    ws_enquiries.column_dimensions["H"].width = 35

    # ── Save to buffer ─────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # ── Log this export for deduplication ───────────────────
    max_c_id = max((c.id for c in contacts), default=last_contact_id)
    max_e_id = max((e.id for e in enquiries), default=last_enquiry_id)

    report_log = WeeklyReportLog(
        week_start=one_week_ago,
        week_end=now,
        record_count=total_records,
        max_contact_id=max_c_id,
        max_enquiry_id=max_e_id,
    )
    db.add(report_log)
    await db.commit()

    logger.info(f"Weekly Excel report generated: {total_records} records ({len(contacts)} contacts, {len(enquiries)} enquiries)")
    return buffer, total_records
